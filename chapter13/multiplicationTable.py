import openpyxl, sys
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet2']

# num = sys.argv[1]
num = int(input())

for rowNum in range(2, num+1):
    sheet.cell(row=rowNum, column=1).value = rowNum-1

for colNum in range(2, num+1):
    sheet.cell(row=1, column=colNum).value = colNum-1

for x in range(2, num+1):
    for y in range(2, num+1):
        sheet.cell(row=y, column=x).value = (x-1)*(y-1)

wb.save('updatedExample.xlsx')